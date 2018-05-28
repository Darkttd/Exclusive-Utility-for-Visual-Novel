import os
import sys
import re
import openpyxl

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print ('usage: python ' + sys.argv[0] + ' scriptName')
        exit()

    wb = openpyxl.load_workbook(sys.argv[1])
    outputScript = None

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        if (ws.cell(row=1,column=1).value is not None):
            outputScript = open('export/' + sheetname + '.rpy', mode='w', encoding='utf8')

            row = 2
            nullCounter = 0

            while nullCounter <= 10:
                eCol = ws.cell(row=row,column=5).value
                dCol = ws.cell(row=row,column=4).value

                isNull = True

                if (eCol is not None):
                    isNull = False
                    print (eCol)
                    outputScript.write("    #" + eCol + "\n")

                if (dCol is not None):
                    isNull = False
                    print (dCol)

                    if (dCol[0:1] == '('):
                        outputScript.write("    \"" + dCol + "\"\n")
                    else:
                        outputScript.write("    아이리 \"" + dCol + "\"\n")

                if (isNull):
                    nullCounter = nullCounter + 1
                    outputScript.write("\n")
                else:
                    nullCounter = 0

                row = row + 1

